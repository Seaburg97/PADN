import os
from pathlib import Path

os.environ.setdefault("MPLCONFIGDIR", "/tmp/matplotlib")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import re
from openpyxl.styles import Font, PatternFill
from sklearn.metrics import (
    cohen_kappa_score,
    confusion_matrix,
    f1_score,
    mean_absolute_error,
    roc_auc_score,
)


def metric_kappa(y_true, pred_label):
    return cohen_kappa_score(
        np.asarray(y_true).astype(int),
        np.asarray(pred_label).astype(int),
        weights="quadratic",
        labels=[0, 1, 2, 3, 4, 5, 6],
    )


def metric_mae(y_true, pred_label):
    return mean_absolute_error(np.asarray(y_true).astype(float), np.asarray(pred_label).astype(float))


def metric_auc(y_binary, prob):
    y_binary = np.asarray(y_binary).astype(int)
    prob = np.asarray(prob).astype(float)
    if len(np.unique(y_binary)) < 2:
        return np.nan
    return roc_auc_score(y_binary, prob)


def calculate_metrics(data, model_prefix):
    true_label = data["true_label"].to_numpy()
    pred_label = data[f"{model_prefix}_pred_label"].to_numpy()
    binary_true = (true_label >= 3).astype(int)
    binary_pred = (pred_label >= 3).astype(int)
    poor_score = data[f"{model_prefix}_prob_poor_outcome"].to_numpy()

    tn, fp, fn, tp = confusion_matrix(binary_true, binary_pred, labels=[0, 1]).ravel()
    return {
        "Kappa": metric_kappa(true_label, pred_label),
        "MAE": metric_mae(true_label, pred_label),
        "AUC": metric_auc(binary_true, poor_score),
        "Sensitivity": tp / (tp + fn) if (tp + fn) else np.nan,
        "Specificity": tn / (tn + fp) if (tn + fp) else np.nan,
        "F1": f1_score(binary_true, binary_pred, zero_division=0),
    }


def bootstrap_metric_values(data, model_prefix, n_bootstrap, rng):
    values = {name: [] for name in METRICS}
    data = data.reset_index(drop=True)
    n_samples = len(data)

    while len(values["AUC"]) < n_bootstrap:
        sampled = data.iloc[rng.integers(0, n_samples, n_samples)]
        binary_true = (sampled["true_label"].to_numpy() >= 3).astype(int)
        if len(np.unique(binary_true)) < 2:
            continue

        metrics = calculate_metrics(sampled, model_prefix)
        for name in METRICS:
            values[name].append(metrics[name])

    return {name: np.asarray(metric_values) for name, metric_values in values.items()}


def table_model_label(model_prefix, model_label):
    if model_prefix == "dl":
        return "DL model"
    return model_label


def table_metric_label(metric_name):
    if metric_name == "F1":
        return "F1 score"
    return metric_name


def grouped_metric_column(metric_name):
    metric_label = table_metric_label(metric_name)
    if metric_name in ("Sensitivity", "Specificity"):
        return f"{metric_label} (95%CI)"
    return f"{metric_label}(95%CI)"


def format_estimate_ci(estimate, ci_low, ci_high):
    return f"{estimate:.3f}({ci_low:.3f}-{ci_high:.3f})"


def summarize_model_with_bootstrap(data, model_prefix, rng):
    point_values = calculate_metrics(data, model_prefix)
    boot_values = bootstrap_metric_values(data, model_prefix, N_BOOTSTRAP, rng)

    summary = {}
    for metric_name in METRICS:
        ci_low, ci_high = np.percentile(boot_values[metric_name], [2.5, 97.5])
        summary[metric_name] = {
            "estimate": float(point_values[metric_name]),
            "ci_low": float(ci_low),
            "ci_high": float(ci_high),
        }
    return summary


def save_grouped_table_excel(grouped_table, long_table):
    GROUPED_TABLE_XLSX.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(GROUPED_TABLE_XLSX, engine="openpyxl") as writer:
        grouped_table.to_excel(writer, index=False, sheet_name="Grouped table")
        long_table.to_excel(writer, index=False, sheet_name="Long format")

        header_fill = PatternFill("solid", fgColor="D9E0E0")
        section_fill = PatternFill("solid", fgColor="E9EEEE")
        bold_font = Font(bold=True)

        worksheet = writer.sheets["Grouped table"]
        for cell in worksheet[1]:
            cell.font = bold_font
            cell.fill = header_fill

        for row_idx in range(2, worksheet.max_row + 1):
            metric_cells_empty = all(
                worksheet.cell(row_idx, col_idx).value in ("", None)
                for col_idx in range(2, worksheet.max_column + 1)
            )
            if metric_cells_empty:
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row_idx, col_idx)
                    cell.font = bold_font
                    cell.fill = section_fill

        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = max(max_length + 2, 14)


def make_grouped_performance_table():
    rng = np.random.default_rng(RANDOM_SEED)
    grouped_columns = [
        "Model",
        "Kappa(95%CI)",
        "MAE(95%CI)",
        "AUC(95%CI)",
        "Sensitivity (95%CI)",
        "Specificity (95%CI)",
        "F1 score(95%CI)",
    ]
    grouped_rows = []
    long_rows = []

    for dataset_key, _ in DATASETS:
        grouped_rows.append({column: "" for column in grouped_columns} | {"Model": dataset_key})
        data = pd.read_csv(INPUT_DIR / f"{dataset_key}_stacking_predictions.csv")

        for model_prefix, model_label, _ in MODELS:
            display_label = table_model_label(model_prefix, model_label)
            summary = summarize_model_with_bootstrap(data, model_prefix, rng)

            grouped_row = {"Model": display_label}
            long_row = {"dataset": dataset_key, "model": display_label}
            for metric_name in METRICS:
                metric_label = table_metric_label(metric_name)
                estimate = summary[metric_name]["estimate"]
                ci_low = summary[metric_name]["ci_low"]
                ci_high = summary[metric_name]["ci_high"]
                grouped_row[grouped_metric_column(metric_name)] = format_estimate_ci(estimate, ci_low, ci_high)
                long_row[metric_label] = estimate
                long_row[f"{metric_label}_ci_low"] = ci_low
                long_row[f"{metric_label}_ci_high"] = ci_high

            grouped_rows.append(grouped_row)
            long_rows.append(long_row)

    grouped_table = pd.DataFrame(grouped_rows, columns=grouped_columns)
    long_table = pd.DataFrame(long_rows)

    GROUPED_TABLE_CSV.parent.mkdir(parents=True, exist_ok=True)
    grouped_table.to_csv(GROUPED_TABLE_CSV, index=False, encoding="utf-8-sig")
    save_grouped_table_excel(grouped_table, long_table)
    return grouped_table, long_table


def draw_panel(ax, dataset_name, panel_label, data, rng):
    x_positions = np.arange(len(METRICS))
    offsets = np.linspace(-0.24, 0.24, len(MODELS))

    for model_index, (model_prefix, model_label, color) in enumerate(MODELS):
        boot_values = bootstrap_metric_values(data, model_prefix, N_BOOTSTRAP, rng)
        point_values = calculate_metrics(data, model_prefix)

        for metric_index, metric_name in enumerate(METRICS):
            x_center = x_positions[metric_index] + offsets[model_index]
            jitter = rng.normal(0, 0.025, N_BOOTSTRAP)
            y_values = boot_values[metric_name]

            ax.scatter(
                np.full(N_BOOTSTRAP, x_center) + jitter,
                y_values,
                s=5,
                color=color,
                alpha=0.32,
                linewidths=0,
            )

            ci_low, ci_high = np.percentile(y_values, [2.5, 97.5])
            ax.vlines(x_center, ci_low, ci_high, color=color, linewidth=2.0, alpha=0.95)
            ax.scatter(
                [x_center],
                [point_values[metric_name]],
                s=38,
                color=color,
                edgecolor="black",
                linewidth=0.5,
                zorder=4,
            )

    ax.axvline(1.5, color="black", linewidth=1.2)
    ax.set_title(f"Performance Metrics - {dataset_name}", fontsize=13)
    ax.set_xticks(x_positions)
    ax.set_xticklabels(METRICS, fontsize=10)
    ax.set_ylabel("Value", fontsize=12)
    ax.set_ylim(Y_LIMIT)
    ax.grid(axis="y", linestyle="--", linewidth=0.5, alpha=0.35)
    ax.grid(axis="x", linestyle=":", linewidth=0.5, alpha=0.25)
    ax.text(
        -0.065,
        1.005,
        panel_label,
        transform=ax.transAxes,
        fontsize=14,
        va="top",
        ha="left",
    )


def draw_integrated_panel(ax, dataset_key, dataset_name, panel_label, data, comparison_df, rng):
    display_metrics = ["Kappa", "MAE", "AUC", "Sensitivity", "Specificity", "F1"]
    comparison_metrics = ["Kappa", "AUC"]
    x_positions = np.arange(len(display_metrics))
    offsets = np.linspace(-0.22, 0.22, len(MODELS))

    for model_index, (model_prefix, model_label, color) in enumerate(MODELS):
        boot_values = bootstrap_metric_values(data, model_prefix, N_BOOTSTRAP, rng)
        for metric_index, metric_name in enumerate(display_metrics):
            x_center = x_positions[metric_index] + offsets[model_index]
            y_values = boot_values[metric_name]
            jitter = rng.normal(0, 0.018, N_BOOTSTRAP)
            ax.scatter(
                np.full(N_BOOTSTRAP, x_center) + jitter,
                y_values,
                s=8,
                color=color,
                alpha=0.20,
                linewidths=0,
                zorder=1,
            )
            ax.hlines(
                np.median(y_values),
                x_center - 0.035,
                x_center + 0.035,
                color="black",
                linewidth=0.7,
                zorder=4,
            )

    comparison_labels = {}
    for metric_name in comparison_metrics:
        p_fusion_padn, _ = comparison_p_value(comparison_df, dataset_key, metric_name, "Fusion vs DL")
        p_clinical, _ = comparison_p_value(comparison_df, dataset_key, metric_name, "Clinical vs DL")
        p_fusion_clinical, _ = comparison_p_value(
            comparison_df, dataset_key, metric_name, "Fusion vs Clinical"
        )
        comparison_labels[metric_name] = (
            f"Fusion-PADN {format_p_value(p_fusion_padn)}\n"
            f"PADN-Clinical {format_p_value(p_clinical)}\n"
            f"Fusion-Clinical {format_p_value(p_fusion_clinical)}"
        )

    ax.axvline(1.5, color="black", linewidth=1.2)
    ax.set_title(f"Performance Metrics - {dataset_name}", fontsize=12, fontweight="bold")
    ax.set_xticks(x_positions)
    ax.set_xticklabels(display_metrics, fontsize=10)
    box_left_positions = {"Kappa": -0.48, "AUC": 1.58}
    for metric_name in comparison_metrics:
        x_pos = box_left_positions[metric_name]
        ax.text(
            x_pos,
            1.78,
            comparison_labels[metric_name],
            fontsize=7.0,
            fontweight="normal",
            linespacing=1.08,
            ha="left",
            va="top",
            clip_on=True,
            bbox={
                "boxstyle": "round,pad=0.22",
                "facecolor": "white",
                "edgecolor": "#bfbfbf",
                "linewidth": 0.7,
                "alpha": 0.92,
            },
        )
    ax.set_ylabel("Value", fontsize=12, fontweight="normal")
    ax.set_ylim(0.0, 1.86)
    ax.tick_params(axis="y", labelsize=10, width=1.3)
    ax.tick_params(axis="x", width=1.3)
    for tick_label in ax.get_yticklabels():
        tick_label.set_fontweight("normal")
    for spine in ax.spines.values():
        spine.set_linewidth(1.3)
    ax.grid(axis="y", linestyle="--", linewidth=0.5, alpha=0.35)
    ax.grid(axis="x", linestyle=":", linewidth=0.45, alpha=0.18)
    handles = [
        plt.Line2D(
            [0],
            [0],
            marker="s",
            color="none",
            markerfacecolor=color,
            markeredgecolor=color,
            markersize=8,
            label=label,
        )
        for _, label, color in MODELS
    ]
    ax.legend(handles=handles, loc="upper right", fontsize=8.5, frameon=True, framealpha=0.92)


def make_integrated_figure():
    rng = np.random.default_rng(RANDOM_SEED + 2000)
    comparison_df = pd.read_csv(COMPARISON_CSV)
    fig, axes = plt.subplots(2, 2, figsize=INTEGRATED_FIGSIZE, dpi=DPI)
    axes = axes.ravel()

    for ax, (dataset_key, title_name), panel_label in zip(axes, DATASETS, PANEL_LABELS):
        data = pd.read_csv(INPUT_DIR / f"{dataset_key}_stacking_predictions.csv")
        draw_integrated_panel(ax, dataset_key, title_name, panel_label, data, comparison_df, rng)

    fig.tight_layout(h_pad=5.8, w_pad=2.2)
    INTEGRATED_OUTPUT_PNG.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(INTEGRATED_OUTPUT_PNG, bbox_inches="tight")
    fig.savefig(INTEGRATED_OUTPUT_PDF, bbox_inches="tight")
    plt.close(fig)


def make_figure():
    rng = np.random.default_rng(RANDOM_SEED)
    plt.rcParams.update(
        {
            "font.size": 9,
            "axes.titlesize": 10,
            "axes.labelsize": 10,
            "xtick.labelsize": 9,
            "ytick.labelsize": 9,
            "legend.fontsize": 7,
        }
    )
    fig, axes = plt.subplots(2, 2, figsize=FIGSIZE, dpi=DPI)
    axes = axes.ravel()

    for ax, (dataset_key, title_name), panel_label in zip(axes, DATASETS, PANEL_LABELS):
        file_path = INPUT_DIR / f"{dataset_key}_stacking_predictions.csv"
        data = pd.read_csv(file_path)
        draw_panel(ax, title_name, panel_label, data, rng)

    handles = [
        plt.Line2D(
            [0],
            [0],
            marker="s",
            color="none",
            markerfacecolor=color,
            markeredgecolor=color,
            markersize=8,
            label=label,
        )
        for _, label, color in MODELS
    ]

    for ax in axes:
        ax.legend(handles=handles, loc="upper right", fontsize=8, frameon=True, framealpha=0.95)

    fig.tight_layout()
    OUTPUT_PNG.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(OUTPUT_PNG, bbox_inches="tight")
    fig.savefig(OUTPUT_PDF, bbox_inches="tight")
    plt.close(fig)


def parse_estimate_ci(text):
    match = re.match(r"([0-9.]+)\(([0-9.]+)-([0-9.]+)\)", str(text))
    if not match:
        return np.nan, np.nan, np.nan
    return tuple(float(item) for item in match.groups())


def format_p_value(p_value):
    if pd.isna(p_value):
        return "P=NA"
    if p_value < 0.001:
        return "P<0.001"
    return f"P={p_value:.3f}"


def comparison_p_value(comparison_df, dataset_key, metric, comparison):
    row = comparison_df[
        (comparison_df["dataset"] == dataset_key)
        & (comparison_df["metric"] == metric)
        & (comparison_df["comparison"] == comparison)
    ]
    if row.empty:
        return np.nan, ""
    return float(row["p_value"].iloc[0]), str(row["test_method"].iloc[0])


def add_pvalue_bracket(ax, x1, x2, y, text, color="#333333"):
    height = 0.015
    ax.plot([x1, x1, x2, x2], [y, y + height, y + height, y], color=color, linewidth=0.8)
    ax.text((x1 + x2) / 2, y + height + 0.003, text, ha="center", va="bottom", fontsize=6, color=color)


def make_metric_comparison_figure():
    table = pd.read_excel(GROUPED_TABLE_XLSX, sheet_name="Grouped table")
    comparison_df = pd.read_csv(COMPARISON_CSV)
    dataset_keys = [item[0] for item in DATASETS]
    dataset_labels = [item[1] for item in DATASETS]
    model_labels = [item[1] for item in MODELS]
    model_colors = [item[2] for item in MODELS]
    metrics = ["Kappa", "AUC"]
    metric_columns = {
        "Kappa": "Kappa(95%CI)",
        "AUC": "AUC(95%CI)",
    }
    fig, axes = plt.subplots(1, 2, figsize=COMPARISON_FIGSIZE, dpi=DPI)
    bar_width = 0.22
    group_x = np.arange(len(dataset_keys))
    offsets = np.linspace(-bar_width, bar_width, len(model_labels))

    for ax, metric in zip(axes, metrics):
        estimates_by_model = {label: [] for label in model_labels}
        lows_by_model = {label: [] for label in model_labels}
        highs_by_model = {label: [] for label in model_labels}
        xtick_labels = []

        for dataset_key in dataset_keys:
            start_idx = table.index[table["Model"] == dataset_key][0]
            block = table.iloc[start_idx + 1 : start_idx + 1 + len(model_labels)]
            for model_label in model_labels:
                source_model_label = "DL model" if model_label == "PADN model" else model_label
                value_text = block.loc[block["Model"] == source_model_label, metric_columns[metric]].iloc[0]
                estimate, low, high = parse_estimate_ci(value_text)
                estimates_by_model[model_label].append(estimate)
                lows_by_model[model_label].append(low)
                highs_by_model[model_label].append(high)

        for model_index, (model_label, color) in enumerate(zip(model_labels, model_colors)):
            estimates = np.asarray(estimates_by_model[model_label])
            lows = np.asarray(lows_by_model[model_label])
            highs = np.asarray(highs_by_model[model_label])
            yerr = np.vstack([estimates - lows, highs - estimates])
            ax.bar(
                group_x + offsets[model_index],
                estimates,
                width=bar_width,
                color=color,
                alpha=0.86,
                edgecolor="black",
                linewidth=0.4,
                label=model_label,
                yerr=yerr,
                capsize=2.5,
                error_kw={"linewidth": 0.8},
            )

        for dataset_index, dataset_key in enumerate(dataset_keys):
            p_fusion_padn, _ = comparison_p_value(comparison_df, dataset_key, metric, "Fusion vs DL")
            p_clinical, _ = comparison_p_value(comparison_df, dataset_key, metric, "Clinical vs DL")
            p_fusion_clinical, _ = comparison_p_value(comparison_df, dataset_key, metric, "Fusion vs Clinical")
            xtick_labels.append(
                f"{dataset_labels[dataset_index]}\n"
                f"Fusion-PADN {format_p_value(p_fusion_padn)}\n"
                f"PADN-Clinical {format_p_value(p_clinical)}\n"
                f"Fusion-Clinical {format_p_value(p_fusion_clinical)}"
            )

        ax.set_title(f"{metric} comparison")
        ax.set_xticks(group_x)
        ax.set_xticklabels(xtick_labels, rotation=0, ha="center", fontsize=5.8, linespacing=1.08)
        ax.set_ylabel("Value")
        ax.set_ylim(COMPARISON_Y_LIMITS[metric])
        ax.grid(axis="y", linestyle="--", linewidth=0.5, alpha=0.35)

    handles, labels = axes[0].get_legend_handles_labels()
    fig.legend(handles, labels, loc="upper center", ncol=3, frameon=False)
    fig.tight_layout(rect=[0, 0.03, 1, 0.93])
    COMPARISON_OUTPUT_PNG.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(COMPARISON_OUTPUT_PNG, bbox_inches="tight")
    fig.savefig(COMPARISON_OUTPUT_PDF, bbox_inches="tight")
    plt.close(fig)


if __name__ == "__main__":
    BASE_DIR = Path(__file__).resolve().parent
    INPUT_DIR = BASE_DIR / ".." / "ClinicalModelAndFusionModel" / "outputs"
    OUTPUT_DIR = BASE_DIR / "outputs"
    OUTPUT_PNG = OUTPUT_DIR / "model_performance_bootstrap_plot.png"
    OUTPUT_PDF = OUTPUT_DIR / "model_performance_bootstrap_plot.pdf"
    GROUPED_TABLE_CSV = OUTPUT_DIR / "paper_table_model_performance_grouped_1000boot.csv"
    GROUPED_TABLE_XLSX = OUTPUT_DIR / "paper_table_model_performance_grouped_1000boot.xlsx"
    COMPARISON_CSV = INPUT_DIR / "model_comparison.csv"
    COMPARISON_OUTPUT_PNG = OUTPUT_DIR / "model_performance_kappa_mae_auc_comparison.png"
    COMPARISON_OUTPUT_PDF = OUTPUT_DIR / "model_performance_kappa_mae_auc_comparison.pdf"
    INTEGRATED_OUTPUT_PNG = OUTPUT_DIR / "model_performance_integrated_core_metrics.png"
    INTEGRATED_OUTPUT_PDF = OUTPUT_DIR / "model_performance_integrated_core_metrics.pdf"
    DATASETS = [
        ("Test-Combined", "Test-Combined"),
        ("efy", "SAMU-R"),
        ("ay2", "FAMU-P"),
        ("th", "TH-R"),
    ]
    MODELS = [
        ("dl", "PADN model", "#3B82B8"),
        ("clinical", "Clinical model", "#D95F5F"),
        ("fused", "Fusion model", "#8E6BBE"),
    ]
    METRICS = ["Kappa", "MAE", "AUC", "Sensitivity", "Specificity", "F1"]
    PANEL_LABELS = ["a", "b", "c", "d"]

    N_BOOTSTRAP = 1000
    RANDOM_SEED = 20260515
    FIGSIZE = (12.5, 7.2)
    DPI = 180
    Y_LIMIT = (0.0, 1.65)
    COMPARISON_FIGSIZE = (10.8, 5.4)
    COMPARISON_Y_LIMITS = {
        "Kappa": (0.42, 1.05),
        "AUC": (0.82, 1.04),
    }
    INTEGRATED_FIGSIZE = (12.8, 7.4)
    INTEGRATED_Y_LIMIT = (0.35, 1.58)

    grouped_table, long_table = make_grouped_performance_table()
    print(f"Saved: {GROUPED_TABLE_CSV}")
    print(f"Saved: {GROUPED_TABLE_XLSX}")
    print(f"Rows: grouped={len(grouped_table)}, long={len(long_table)}")

    make_metric_comparison_figure()
    print(f"Saved: {COMPARISON_OUTPUT_PNG}")
    print(f"Saved: {COMPARISON_OUTPUT_PDF}")

    make_integrated_figure()
    print(f"Saved: {INTEGRATED_OUTPUT_PNG}")
    print(f"Saved: {INTEGRATED_OUTPUT_PDF}")
