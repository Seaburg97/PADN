from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill
from scipy.stats import kruskal


def excel_col_to_index(col_name):
    col_name = col_name.strip().upper()
    index = 0
    for char in col_name:
        if not ("A" <= char <= "Z"):
            raise ValueError(f"Invalid Excel column name: {col_name}")
        index = index * 26 + ord(char) - ord("A") + 1
    return index - 1


def read_csvs(paths):
    frames = []
    for path in paths:
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")
        frames.append(pd.read_csv(path))
    if not frames:
        raise ValueError("No CSV files were provided")
    return pd.concat(frames, ignore_index=True)


def get_columns_by_excel_range(df, start_col, end_col):
    start_idx = excel_col_to_index(start_col)
    end_idx = excel_col_to_index(end_col)
    if start_idx < 0 or end_idx >= len(df.columns) or start_idx > end_idx:
        raise ValueError(
            f"Column range {start_col}:{end_col} is outside the table width; the table has {len(df.columns)} columns"
        )
    return list(df.columns[start_idx : end_idx + 1])


def numeric_values(series):
    return pd.to_numeric(series, errors="coerce").dropna()


def clean_level(value):
    if pd.isna(value):
        return "Missing"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def format_number(value, digits=2):
    if pd.isna(value):
        return "NA"
    text = f"{value:.{digits}f}"
    return text.rstrip("0").rstrip(".")


def format_percent(value):
    if pd.isna(value):
        return "NA"
    return f"{value:.2f}"


def format_mean_sd(series):
    values = numeric_values(series)
    if values.empty:
        return "NA"
    mean = values.mean()
    sd = values.std(ddof=1) if len(values) > 1 else 0
    return f"{format_number(mean)} ± {format_number(sd)}"


def format_median_iqr(series):
    values = numeric_values(series)
    if values.empty:
        return "NA"
    q1 = values.quantile(0.25)
    median = values.quantile(0.50)
    q3 = values.quantile(0.75)
    return f"{format_number(median)} ({format_number(q1)}, {format_number(q3)})"


def kruskal_p_value(groups, variable):
    values_by_group = []
    for df in groups.values():
        values = numeric_values(df[variable])
        if not values.empty:
            values_by_group.append(values)
    if len(values_by_group) < 2:
        return np.nan
    try:
        return kruskal(*values_by_group).pvalue
    except ValueError:
        return np.nan


def format_continuous(groups, variable, alpha):
    p_value = kruskal_p_value(groups, variable)
    use_median_iqr = pd.notna(p_value) and p_value < alpha
    formatter = format_median_iqr if use_median_iqr else format_mean_sd
    return {group_name: formatter(df[variable]) for group_name, df in groups.items()}


def sorted_levels(series):
    values = series.dropna().unique()
    try:
        return sorted(values)
    except TypeError:
        return sorted(values, key=lambda value: str(value))


def format_categorical_group(series, levels):
    valid = series.dropna()
    total = len(valid)
    if total == 0:
        return {level: "NA" for level in levels}

    counts = valid.value_counts(dropna=False)
    values = {}
    for level in levels:
        count = counts.get(level, 0)
        percent = count / total * 100
        values[level] = f"{int(count)} ({format_percent(percent)}%)"
    return values


def format_categorical(groups, variable, levels):
    return {
        group_name: format_categorical_group(df[variable], levels)
        for group_name, df in groups.items()
    }


def build_summary_table(groups, variables, alpha=0.05, continuous_variables=None):
    continuous_variables = set(continuous_variables or [])
    group_names = list(groups.keys())
    rows = []
    for variable in variables:
        combined = pd.concat([df[variable] for df in groups.values()], ignore_index=True)
        if variable in continuous_variables:
            values = format_continuous(groups, variable, alpha)
            row = {"Variable": variable}
            row.update(values)
            rows.append(row)
        else:
            levels = sorted_levels(combined)
            values = format_categorical(groups, variable, levels)
            rows.append({"Variable": variable, **{group_name: "" for group_name in groups}})
            for level in levels:
                row = {"Variable": clean_level(level)}
                for group_name in groups:
                    row[group_name] = values[group_name][level]
                rows.append(row)
    table = pd.DataFrame(rows, columns=["Variable", *group_names])
    table = table.rename(
        columns={group_name: f"{group_name} (n={len(df)})" for group_name, df in groups.items()}
    )
    return table


def save_excel(table, output_xlsx):
    output_xlsx = Path(output_xlsx)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        table.to_excel(writer, index=False, sheet_name="Clinical variables")
        worksheet = writer.sheets["Clinical variables"]
        header_fill = PatternFill("solid", fgColor="D9E0E0")
        section_fill = PatternFill("solid", fgColor="E9EEEE")
        bold_font = Font(bold=True)

        for cell in worksheet[1]:
            cell.font = bold_font
            cell.fill = header_fill

        for row_idx in range(2, worksheet.max_row + 1):
            if all(worksheet.cell(row_idx, col_idx).value in ("", None) for col_idx in range(2, worksheet.max_column + 1)):
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row_idx, col_idx)
                    cell.font = bold_font
                    cell.fill = section_fill

        worksheet.freeze_panes = "A2"
        worksheet.column_dimensions["A"].width = 24
        for column_cells in worksheet.iter_cols(min_col=2, max_col=worksheet.max_column):
            worksheet.column_dimensions[column_cells[0].column_letter].width = 24


def main():
    development_frames = {
        group_name: read_csvs([csv_path])
        for group_name, csv_path in DEVELOPMENT_CSV_PATHS.items()
    }
    development_df = pd.concat(development_frames.values(), ignore_index=True)
    test_efy_df = read_csvs([EFY_CSV_PATH])
    test_ay2_df = read_csvs([AY2_CSV_PATH])
    test_th_df = read_csvs([TH_CSV_PATH])

    variables = get_columns_by_excel_range(development_df, START_EXCEL_COLUMN, END_EXCEL_COLUMN)
    groups = {
        "Development": development_df,
        "efy": test_efy_df,
        "ay2": test_ay2_df,
        "th": test_th_df,
    }
    split_groups = {
        **development_frames,
        "efy": test_efy_df,
        "ay2": test_ay2_df,
        "th": test_th_df,
    }
    table = build_summary_table(groups, variables, KW_ALPHA, CONTINUOUS_VARIABLES)
    split_table = build_summary_table(split_groups, variables, KW_ALPHA, CONTINUOUS_VARIABLES)
    save_excel(table, OUTPUT_XLSX_PATH)
    save_excel(split_table, OUTPUT_SPLIT_XLSX_PATH)

    print(f"Excel saved: {OUTPUT_XLSX_PATH}")
    print(f"Development-split Excel saved: {OUTPUT_SPLIT_XLSX_PATH}")
    print(table)


BASE_DIR = Path(__file__).resolve().parent
FEATURE_DIR = BASE_DIR / "data" / "features"
OUTPUT_DIR = BASE_DIR / "outputs"

DEVELOPMENT_CSV_PATHS = {
    "aq": FEATURE_DIR / "featuresaq.csv",
    "ay": FEATURE_DIR / "featuresay.csv",
    "fy": FEATURE_DIR / "featuresfy.csv",
    "tl": FEATURE_DIR / "featurestl.csv",
    "yjs": FEATURE_DIR / "featuresyjs.csv",
}

EFY_CSV_PATH = FEATURE_DIR / "featuresefy.csv"
AY2_CSV_PATH = FEATURE_DIR / "featuresay2.csv"
TH_CSV_PATH = FEATURE_DIR / "featuresth.csv"

START_EXCEL_COLUMN = "CS"
END_EXCEL_COLUMN = "DG"
KW_ALPHA = 0.05
CONTINUOUS_VARIABLES = [
    "GCS_score",
    "Age",
    "Size",
]

OUTPUT_XLSX_PATH = OUTPUT_DIR / "clinical_variable_summary.xlsx"
OUTPUT_SPLIT_XLSX_PATH = OUTPUT_DIR / "clinical_variable_summary_by_development_center.xlsx"


if __name__ == "__main__":
    main()
